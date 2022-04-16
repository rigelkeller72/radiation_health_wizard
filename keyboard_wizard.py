# Automated Keyboard and Mouse Demo
# Rigel Keller
# Best example inside either a google sheets or excel spreadsheet

from pynput.keyboard import Key, Controller as KeyboardController
from pynput.mouse import Button, Controller as MouseController
import time
import openpyxl

path = "random_data.xlsx"
time.sleep(1)

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
# max_row = sheet_obj.max_row
max_row = 10
big_lst = []


# Loop will print all columns name
for x in range(1, max_row + 1):
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=x, column=i)
        box = cell_obj.value
        if type(box) == int:
            box = str(box)
        big_lst.append(box)
        time.sleep(0.1)

# Converting list into list of Tuples
lst_tuple = [x for x in zip(*[iter(big_lst)]*5)]

# Keyboard and mouse setup
keyboard = KeyboardController()
mouse = MouseController()
t = Key.tab
e = Key.enter

type_speed = 0.2
time.sleep(3)


def myfunction(first_name, last_name, age, ssn, dob):
    # Set pointer position
    mouse.position = (180, 330)
    time.sleep(type_speed)

    # Press and release
    mouse.press(Button.left)
    mouse.release(Button.left)
    time.sleep(type_speed)

    for repeat in range(4):
        keyboard.press(t)
        keyboard.release(t)
        time.sleep(type_speed)

    for char in first_name:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)

    for char in last_name:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)

    for char in age:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)

    for char in ssn:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)

    for char in dob:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)
    keyboard.press(e)
    keyboard.release(e)
    time.sleep(type_speed)

    # ************************second window input*******************
    # Set pointer position
    mouse.position = (180, 330)
    time.sleep(type_speed)

    # Press and release
    mouse.press(Button.left)
    mouse.release(Button.left)
    time.sleep(type_speed)

    for repeat2 in range(4):
        keyboard.press(t)
        keyboard.release(t)
        time.sleep(type_speed)

    for char in first_name:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)

    for char in last_name:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)

    for char in age:
        keyboard.press(char)
        keyboard.release(char)
        time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)
    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)
    keyboard.press(e)
    keyboard.release(e)
    time.sleep(type_speed)

    # ************** third window ***************************
    # Set pointer position
    mouse.position = (180, 330)
    time.sleep(type_speed)

    # Press and release
    mouse.press(Button.left)
    mouse.release(Button.left)
    time.sleep(type_speed)

    keyboard.press(t)
    keyboard.release(t)
    time.sleep(type_speed)
    keyboard.press(e)
    keyboard.release(e)
    time.sleep(type_speed)


# Start going through patient data
for index, tuples in enumerate(lst_tuple):
    myfunction(tuples[0], tuples[1], tuples[2], tuples[3], tuples[4])
